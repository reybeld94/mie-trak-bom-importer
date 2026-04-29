from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.models.bom_line import BOMLine

_HEADER_PART_NUMBER = {
    "part",
    "part number",
    "part #",
    "pn",
    "p/n",
    "component",
    "component part",
    "component part number",
    "item",
    "item number",
}

_HEADER_DESCRIPTION = {
    "description",
    "desc",
    "item description",
    "component description",
}

_HEADER_QUANTITY = {
    "qty",
    "quantity",
    "req qty",
    "required qty",
    "required quantity",
}

_HEADER_UOM = {
    "uom",
    "unit",
    "unit of measure",
    "units",
}

_REQUIRED_FIELDS = ("part_number", "description", "quantity", "unit_of_measure")


class PackingListExcelParser:
    def parse_xlsx(self, file_path: str | Path) -> list[BOMLine]:
        path = Path(file_path)
        workbook = load_workbook(filename=path, data_only=True, read_only=True)

        parsed_lines: list[BOMLine] = []

        for worksheet in workbook.worksheets:
            header_row_idx, header_columns = self._detect_header(worksheet)
            if header_row_idx is None or header_columns is None:
                continue

            for row_number, row_values in enumerate(
                worksheet.iter_rows(min_row=header_row_idx + 1, values_only=True),
                start=header_row_idx + 1,
            ):
                values_by_field = {
                    field_name: self._cell_to_string(
                        row_values[column_idx] if column_idx < len(row_values) else None
                    )
                    for field_name, column_idx in header_columns.items()
                }

                part_number = values_by_field["part_number"]
                description = values_by_field["description"]
                quantity_value = self._normalize_quantity(
                    row_values[header_columns["quantity"]]
                    if header_columns["quantity"] < len(row_values)
                    else None
                )
                unit_of_measure = values_by_field["unit_of_measure"]

                is_blank_row = all(value == "" for value in values_by_field.values())
                if is_blank_row:
                    continue

                if part_number == "" and quantity_value is None:
                    continue

                parsed_lines.append(
                    BOMLine(
                        source_file=path.name,
                        source_sheet=worksheet.title,
                        source_row=row_number,
                        component_part_number=part_number,
                        description=description,
                        quantity=quantity_value if quantity_value is not None else 0.0,
                        unit_of_measure=unit_of_measure,
                        notes="",
                        status="",
                        message="",
                    )
                )

        workbook.close()
        return parsed_lines

    def _detect_header(self, worksheet: Any) -> tuple[int | None, dict[str, int] | None]:
        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            normalized = [self._normalize_header_value(cell) for cell in row]
            mapping = self._map_columns(normalized)
            if all(field in mapping for field in _REQUIRED_FIELDS):
                return row_idx, mapping

            if row_idx >= 50:
                break

        return None, None

    def _map_columns(self, normalized_cells: list[str]) -> dict[str, int]:
        mapping: dict[str, int] = {}
        for index, cell_value in enumerate(normalized_cells):
            if not cell_value:
                continue

            if "part_number" not in mapping and cell_value in _HEADER_PART_NUMBER:
                mapping["part_number"] = index
                continue

            if "description" not in mapping and cell_value in _HEADER_DESCRIPTION:
                mapping["description"] = index
                continue

            if "quantity" not in mapping and cell_value in _HEADER_QUANTITY:
                mapping["quantity"] = index
                continue

            if "unit_of_measure" not in mapping and cell_value in _HEADER_UOM:
                mapping["unit_of_measure"] = index

        return mapping

    @staticmethod
    def _normalize_header_value(value: Any) -> str:
        if value is None:
            return ""

        normalized = str(value).strip().lower()
        normalized = normalized.replace("_", " ")
        return " ".join(normalized.split())

    @staticmethod
    def _cell_to_string(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _normalize_quantity(value: Any) -> float | None:
        if value is None:
            return None

        if isinstance(value, (int, float)):
            return float(value)

        normalized = str(value).strip().replace(",", "")
        if normalized == "":
            return None

        try:
            return float(normalized)
        except ValueError:
            return None
